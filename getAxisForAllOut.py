#!/usr/bin/python2.7
# -*- coding: utf-8 -*-
import sys
import os

from math import cos,sin,radians,sqrt,fabs,acos,pi,degrees,floor
import openpyxl
from openpyxl import load_workbook
import xml.etree.ElementTree
from geographiclib.geodesic import Geodesic
from geopy.distance import VincentyDistance
geod = Geodesic.WGS84


axeRadius = [0.0,0.2,0.4,0.6,0.8,1.0,2.0,3.0,4.0,5.0,6.0,7.0,8.0,9.0,10.0,12.0,14.0,16.0,18.0,20.0,
            22.0,24.0,26.0,28.0,30.0,32.0,34.0,36.0,38.0,40.0,42.0,44.0,46.0,48.0,50.0,
	    52.0,54.0,56.0,58.0,60.0]
tol = 1000
axePoints = []

srcPosLon = 33.5403694792
srcPosLat = 36.1449943051

def angleOf1(dX,dY):
	dFi = 0.0
	dR = sqrt(fabs(dX*dX+dY*dY));
	if (fabs(dR) > 0):
		dFi  = acos(dX/dR);
	if (dX <= 0 and dY < 0):
		dFi  = 2.0*pi - dFi;
	if (dX >  0 and dY < 0):
		dFi  = 2.0*pi - dFi;
	return dFi;
tab = 2
coordinates = [[[geod.Direct(srcPosLat, srcPosLon, 360.0-(degrees(angleOf1(r1*1000.0*cos(radians(float(phi)/10.0)),-r1*1000.0*sin(radians(float(phi)/10.0))))-90.0), r1*1000.0)['lon2'],\
                 geod.Direct(srcPosLat, srcPosLon, 360.0-(degrees(angleOf1(r1*1000.0*cos(radians(float(phi)/10.0)),-r1*1000.0*sin(radians(float(phi)/10.0))))-90.0), r1*1000.0)['lat2']] for phi in range(0,3600)] for r1 in axeRadius]

#print coordinates
			
def find_element_in_list(element,list_element):
        try:
		index_element=list_element.index(element)
		return index_element
	except ValueError:
		return -1 

def readLineF(line):
	lst = line.rstrip()
	vals = filter(None,lst.split(" "))
	return [float(x) for x in vals]

def readLineI(line):
	lst = line.rstrip()
	vals = filter(None,lst.split(" "))
	return [int(x) for x in vals]

def cleanBlock(block):
	for i in range (len(block)-1, -1, -1):
		if(block[i].strip().strip("\t") == ""):
			del block[i]
	return block

class GeoGrid:
	def __init__(self,gridFunction):
		self.lonmin = 0.0
		self.lonmax = 0.0
		self.latmin = 0.0
		self.latmax = 0.0
		self.countx = 0.0
		self.county = 0.0
		self.fmin   = 0.0
		self.fmax   = 0.0
		
		self.calcFunctionId = -1
		self.tipObl = -1
		self.isotop = ""
		self.phchForm = ""
		self.organ = ""
		self.age = ""
		self.timeEnd = 0.0
		
		self.data   = []

		if "calcFunctionId" not in gridFunction.attrib:
			self.calcFunctionId = -1
		else:
			self.calcFunctionId = int(gridFunction.get("calcFunctionId").strip())
		
		if "tipObl" not in gridFunction.attrib:
			self.tipObl = -1
		else:
			self.tipObl = int(gridFunction.get("tipObl").strip())
		
		if "isotop" not in gridFunction.attrib:
			self.isotop = ""
		else:
			self.isotop = gridFunction.get("isotop").strip()
			
		if "phchForm" not in gridFunction.attrib:
			self.phchForm = ""
		else:
			self.phchForm = gridFunction.get("phchForm").strip()
			
		if "organ" not in gridFunction.attrib:
			self.organ = ""
		else:
			self.organ = gridFunction.get("organ").strip()
			
		if "age" not in gridFunction.attrib:
			self.age = ""
		else:
			self.age = gridFunction.get("age").strip()
			
		if "timeEnd" not in gridFunction.attrib:
			self.timeEnd = 0.0
		else:
			self.timeEnd = float(gridFunction.get("timeEnd").strip())
				
		lst = gridFunction.text.split("\n")
		lst = cleanBlock(lst)

		self.countx = int(lst[0].split(" ")[1])
		self.county = int(lst[1].split(" ")[1])
		self.lonmin = float(lst[2].split(" ")[1])
		self.latmin = float(lst[3].split(" ")[1])
		self.dlon  =  float(lst[4].split(" ")[1])
		self.dlat  =  float(lst[5].split(" ")[1])
		self.lonmax = self.lonmin+self.dlon*(self.countx-1)
		self.latmax = self.latmin+self.dlat*(self.county-1)

		lst = lst[7:]
		self.data   = [0.0]*self.countx*self.county
		for j in range(len(lst)): #количество строк
			lin = readLineF(lst[j])
			for i in range(len(lin)):
				self.data[i*self.county+(self.county-j-1)] = lin[i]
		
		self.fmin   = self.findMin()
		self.fmax   = self.findMax()

		self.printASCIIGRDFile("test.grd")
	
	def findMin(self):
		if(len(self.data) == 0):
			return -1
		ret = self.data[0]
		for v in self.data:
			if(ret > v):
				ret = v
		return ret
		
	def findMax(self):
		if(len(self.data) == 0):
			return -1
		ret = self.data[0]
		for v in self.data:
			if(ret < v):
				ret = v
		return ret			

	def printASCIIGRDFile(self,filename):
		f = open(filename, 'wt')
		f.write("DSAA" + '\r\n')
		f.write(str(self.countx)+"\t"+str(self.county) + '\r\n')
		f.write('%0.12f' % self.lonmin+"\t"+'%0.12f' % self.lonmax + '\r\n')
		f.write('%0.12f' % self.latmin+"\t"+'%0.12f' % self.latmax + '\r\n')
		f.write('%0.12e' % self.findMin()+"\t"+'%0.12e' % self.findMax() + '\r\n')
		for j in range(0,self.county):
			for i in range(0,self.countx):
				f.write('%0.12e' % self.data[i*self.county+j]+'\t')
			f.write('\r\n')
		
		f.close()
		return

	def getValue(self,lon,lat):
		if(lon<self.lonmin or lon>self.lonmax):
			return 0.
		if(lat<self.latmin or lat>self.latmax):
			return 0.
		ci = int(floor((lon-self.lonmin)/self.dlon))
		cj = int(floor((lat-self.latmin)/self.dlat))
		ci1 = 0
		ci2 = 0
                cj1 = 0
		cj2 = 0
		
		if ci == self.countx-1:
			ci2 = ci
			ci1 = ci-1
		else:
			ci1 = ci
			ci2 = ci+1
	    
		if cj == self.county-1:
			cj2 = cj
			cj1 = cj-1
		else:
			cj1 = cj
			cj2 = cj+1
		
		f11 = self.data[ci1*self.county+cj1] #+ -
		f21 = self.data[ci2*self.county+cj1] #+ -

		f12 = self.data[ci1*self.county+cj2] #- +
		f22 = self.data[ci2*self.county+cj2] #- +
		
		
		x1 = self.lonmin+self.dlon*ci1
		x2 = self.lonmin+self.dlon*ci2
		y1 = self.latmin+self.dlat*cj1
		y2 = self.latmin+self.dlat*cj2
		
		fy1=(f21-f11)/(x2-x1)*lon+(f11-(f21-f11)/(x2-x1)*x1)

		fy2=(f22-f12)/(x2-x1)*lon+(f12-(f22-f12)/(x2-x1)*x1)

		v = (fy2-fy1)/(y2-y1)*lat+(fy1-(fy2-fy1)/(y2-y1)*y1)
		return  v
	
	def angleOf(self,dX,dY):
		dFi = 0.0
		dR = sqrt(fabs(dX*dX+dY*dY));
		if (fabs(dR) > 0):
			dFi  = acos(dX/dR);
		if (dX <= 0 and dY < 0):
			dFi  = 2.0*pi - dFi;
		if (dX >  0 and dY < 0):
			dFi  = 2.0*pi - dFi;
		return dFi;
	
	def getPlumeAxisVals(self, rVals):
		res = []
		phiMin = 0.0
		phiMax = 360.0
		dphi = 0.1
		#print self.getValue(srcPosLon,srcPosLat)
		#r = 1000.0
		for r in rVals:
			r1 = r*1000.0
			fv = 0. 
			phi = phiMin
			while phi <= phiMax:
				dx = r1*cos(radians(phi))
				dy = -r1*sin(radians(phi))
				azimut = self.angleOf(dx,dy)
				#print 360.0-(degrees(azimut)-90.0)
				g = geod.Direct(srcPosLat, srcPosLon, 360.0-(degrees(azimut)-90.0), r1)
			
				lon = g['lon2']
				lat = g['lat2']
				tmp = self.getValue(lon,lat)
				if(fv<tmp):
					fv = tmp
				#print lon,lat,phi,self.getValue(lon,lat)
				phi = phi+dphi
			#print r,fv
			res.append(fv)
		
		#for r in range(len
		
		return res
	
	def getPlumeAxisPoints(self, rVals):
		res = []
		phiMin = 0.0
		phiMax = 360.0
		dphi = 0.001
		#print self.getValue(srcPosLon,srcPosLat)
		#r = 1000.0
		for r in rVals:
			point = [0.0,0.0]
			r1 = r*1000.0
			fv = 0. 
			phi = phiMin
			while phi <= phiMax:
				dx = r1*cos(radians(phi))
				dy = -r1*sin(radians(phi))
				azimut = self.angleOf(dx,dy)
				#print 360.0-(degrees(azimut)-90.0)
				g = geod.Direct(srcPosLat, srcPosLon, 360.0-(degrees(azimut)-90.0), r1)
			
				lon = g['lon2']
				lat = g['lat2']
				tmp = self.getValue(lon,lat)
				if(fv<tmp):
					fv = tmp
					point[0] = lon
					point[1] = lat
				#print lon,lat,phi,self.getValue(lon,lat)
				phi = phi+dphi
			#print r,fv
			res.append(point)
		return res
	
	def getPlumeAxisPoints2(self, coordinates):
		res = []

		#print self.getValue(srcPosLon,srcPosLat)
		#r = 1000.0
		for phiArr in coordinates:
			point = [0.0,0.0]
			fv = 0. 
			for phi in phiArr:
				lon = float(str("%.10f" % phi[0]))
				lat = float(str("%.10f" % phi[1]))
				tmp = self.getValue(lon,lat)
				if(fv<tmp):
					fv = tmp
					point[0] = lon
					point[1] = lat
			res.append(point)
		return res
		
	def getPlumeAxisValsForPoints(self, rPoints):
		res = []
		for r in rPoints:
			tmp = self.getValue(r[0],r[1])
			res.append(tmp)
		return res
def compare(ww, name, mypath):
	lstcompare = []	#список имен функционалов
	lst2compare=[]	# Создаем список сортировки
	for j in range(206):
		lstcompare.append(ww.cell(row=1, column=j+2).value)
	print len(lstcompare)
	for i in range(2, len(lstcompare)+2): #+2 потому что так построен входной файл, цикл по строкам
		find=False
		if len(lst2compare)>0 and i<208:
			for k in range(len(lst2compare)):
				if lst2compare[k].count(lstcompare[i-2])!=0:
					find=True
		if find==True:
			continue
		lst2compare.append([])
		lst2compare[len(lst2compare)-1].append(lstcompare[i-2])
		for j in range(2, len(lstcompare)+2): #+2 потому что так построен входной файл, цикл по столбцам
			if i==j:
				continue
			if int(ww.cell(row=i, column=j).value)==1:
				#print "sovpalo"
				lst2compare[len(lst2compare)-1].append(lstcompare[j-2]) #Добавляем элемент в список, если он отсутствует
				
			#else:
				#print  ws.cell(row=i, column=j).value

	f = open(mypath+"/CompareAll/"+name+str(".txt"), 'wt')
	f.write(str(len(lst2compare)) + '\n')
	for i in range(len(lst2compare)):
		print lst2compare[i]
		
		f.write(str(lst2compare[i]) + '\r\n')
	print len(lst2compare)

	f.close()
	return

def eqArr(arr1,arr2):
	if len(arr1) != len(arr2):
		return False
	
	if (len(arr1) == 0 and len(arr2) == 0):
		return True
	
	count = len(arr1)
	
	for i in range(count):
		g = geod.Inverse(arr1[i][1], arr1[i][0], arr2[i][1], arr2[i][0]) #расстояние между ними
		#print arr1[i][1], arr2[i][1], arr1[i][0], arr2[i][0], g['s12']
		if g['s12']>float(tol):
		#if abs(arr1[i][0] - arr2[i][0]) > 1.0e-4 or abs(arr1[i][1] - arr2[i][1]) > 1.0e-4:
			return False
	
	return True
def main():
	mypath=os.path.dirname(os.path.realpath( __file__ ))
	os.chdir(mypath)
	#if(len(sys.argv)==1):
	#	return
	
	global axePoints, name
	lst = ["calcFunctionId","tipObl","isotop","phchForm","organ","age","timeEnd"]
	lst.extend(axeRadius)
	alltime = 366*24*3600/7200
	for ii in range(alltime+1):
		path = str("/home/egor/work/results/"+str(ii*7200)+" s/out.xml")
		print path 
		wb = openpyxl.Workbook()
		ws = wb.worksheets[0]
		ws.title = u'Сравнение осей следа'
		et = xml.etree.ElementTree.parse(str(path))
		root = et.getroot()
		grid = root.find('grid')
		numbers=[]
		gres = []
		gres.append(lst)
		totArray = []
		totArrayNonEq = []
		for gridFunctionId, gridFunction in enumerate(grid.findall('gridFunction')):
			ws.cell(row=1, column=gridFunctionId + 2, value="f"+str(gridFunctionId))
			ws.cell(row=gridFunctionId+2, column=1, value="f"+str(gridFunctionId))
			print str(gridFunctionId)+"/"+str(len(grid.findall('gridFunction')))
			fil = GeoGrid(gridFunction)
			axePoints = fil.getPlumeAxisPoints2(coordinates)
			totArray.append(axePoints)
			del fil
			find = False
		
			for k, arr in enumerate(totArray):
				
				if eqArr(arr, axePoints):
					find = True
					ws.cell(row=gridFunctionId + 2, column=k + 2, value=str("1"))
					ws.cell(row=k + 2, column=gridFunctionId + 2, value=str("1"))   #симметрично заполняем таблицу относительно диагонали

				else:
					ws.cell(row=gridFunctionId + 2, column=k + 2, value=str("0"))
					ws.cell(row=k+2, column=gridFunctionId+2, value=str("0"))   #симметрично заполняем таблицу относительно диагонали
				if k==gridFunctionId:
					ws.cell(row=k+2, column=k+2, value=str("//////////////"))
				
			if find == False:
				numbers.append(gridFunctionId)
				totArrayNonEq.append(axePoints) 
		compare(ws, str(ii*7200), mypath)
		if os.path.isdir('AxisAll')==False:
			os.mkdir('AxisAll')
		if os.path.isdir('CompareAll')==False:
			os.mkdir('CompareAll')
		name=str(mypath + "/CompareAll/Compare_"+str(ii*7200)+".xlsx")
		wb.save(name)
		f = open(mypath+"/AxisAll/axe_"+str(ii*7200)+".dat", 'wt')
		for i in range(len(totArray)):
			f.write(str(i)+';'+str(i)+';')
		f.write('\r\n')
		for i in range(len(totArray)):
			f.write('lon;lat;')
		f.write('\r\n')
		for i in range(len(totArray[0])):
			for j in range(len(totArray)):
				f.write('%0.12f' % totArray[j][i][0] + ';'+'%0.12f' % totArray[j][i][1]+ ';')
			f.write('\r\n')
		f.close()	
		
		f1 = open(mypath+"/AxisAll/axe_"+str(ii*7200)+"_NE.dat", 'wt')
		for i in range(len(totArrayNonEq)):
			f1.write(str(numbers[i])+';'+str(i)+';')
		f1.write('\r\n')
		for i in range(len(totArrayNonEq)):
			f1.write('lon;lat;')
		#f1.write('\r\n')
		#for i in range(len(totArrayNonEq[0])):
		#	for j in range(len(totArrayNonEq)):
		#		f1.write('%0.12f' % totArrayNonEq[j][i][0] + ';'+'%0.12f' % totArrayNonEq[j][i][1]+ ';')
		#	f1.write('\r\n')
		#f1.close()
		print str("            ")+str(ii*7200)+str(" s   done")
		#break
		
	
	print "done!!!!!"
	return

if __name__ == "__main__":
	sys.exit(main())