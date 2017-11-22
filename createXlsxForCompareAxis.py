#!/usr/bin/python2.7
# -*- coding: utf-8 -*-
import sys
import os

from math import cos,sin,radians,sqrt,fabs,acos,pi,degrees,floor

import xml.etree.ElementTree
from geographiclib.geodesic import Geodesic
from geopy.distance import VincentyDistance
geod = Geodesic.WGS84
import numpy as np
import openpyxl
from openpyxl import load_workbook
axeRadius = [0.0,0.2,0.4,0.6,0.8,1.0,2.0,3.0,4.0,5.0,6.0,7.0,8.0,9.0,10.0,12.0,14.0,16.0,18.0,20.0,
			22.0,24.0,26.0,28.0,30.0,32.0,34.0,36.0,38.0,40.0,42.0,44.0,46.0,48.0,50.0,
		52.0,54.0,56.0,58.0,60.0]
tab=2 #ключ заполнения таблицы. Если 1, то заполняется процентами, если 2, то ноликами и единичками, если 3, то количеством точек
pathToAxis="/home/egor/work/Axis/"
axePoints=[]
tol=1000 #Толератность в метрах


def main():
	print len(axeRadius)
	global axePoints
	eq=False
	length=float(len(axeRadius))
	axePoints2=[]
	mypath = os.path.dirname(os.path.realpath(__file__))
	os.chdir(mypath)
	lst=[]
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Сравнение осей следа'
	for filename in os.listdir(str(pathToAxis)):
		lst.append(filename)
	lst.sort()
	for j in range(len(lst)): #перебираем все функционалы(1ый функц. для сравнения)(что сравниваем)
		axePoints = []
		print lst[j]
		#ws['A'+str(j+2)] = format(str(lst[j].split('_')[1].split('.')[0]))
		ws.cell(row=1, column=j + 2, value=str(lst[j].split('_')[1].split('.')[0]))
		ws.cell(row=j+2, column=1, value=str(lst[j].split('_')[1].split('.')[0]))
		f  = open(str(pathToAxis + str(lst[j])))
		string = f.read()
		f.close()
		for i in range(len(axeRadius)): #перебираем все точки оси следа в функционале, заполняем массив для сравнения
			point=[]
			lst2 = string.split("\n")
			point = [float(lst2[i].split(" ")[0].replace('\r','').replace(',', '')), float(lst2[i].split(" ")[1].replace('\r','').replace(',', '')) ]
			axePoints.append(point)
		for k in range(j, len(lst)): #перебира по всем функц. (2ой функционал для сравнения)(с чем сравниваем) от jтого функционал и до конца, чтобы 2 раза не пересчитывать

			if k==j:
				ws.cell(row=k+2, column=k+2, value=str("//////////////"))
				continue
			#print lst[k]
			f2 = open(str(pathToAxis + str(lst[k])))
			string2 = f2.read()
			f2.close()
			n=0
			axePoints2 = []
			for i in range(len(axeRadius)):  #во 2ом функционале сравниваем все точки с первым
				point=[]
				lst2 = string2.split("\n")
				point = [float(lst2[i].split(" ")[0].replace(',', '').replace('\r', '')), float(lst2[i].split(" ")[1].replace(',', '').replace('\r', ''))]
				axePoints2.append(point)
				# g = geod.Inverse(lat1, lon1, lat2, lon2) distance between lat1;lon1 and lat2, lon2
				#Сравнение с толерантностью
				g = geod.Inverse(float(axePoints[i][1]), float(axePoints[i][0]), float(axePoints2[i][1]), float(axePoints2[i][0]))#расстояние между ними
				if g['s12']>float(tol):
				#Вариант сравнения по координатам
				#if abs(axePoints[i][1] - axePoints2[i][1]) > 1.0e-2 or abs(axePoints[i][0] - axePoints2[i][0]) > 1.0e-2:
						eq=False
				else:
						eq=True
				if eq:
					n=n+1
			if tab == 1:
				var="per"
				ws.cell(row=j + 2, column=k + 2, value="{:.1f}".format(float(n / 40.0 * 100.0)))
				ws.cell(row=k + 2, column=j + 2, value="{:.1f}".format(float(n / 40.0 * 100.0)))
			elif tab == 2:
				var="01"
				if n==40: #симметрично заполняем таблицу относительно диагонали
					ws.cell(row=j + 2, column=k + 2, value=str("1"))
					ws.cell(row=k+2, column=j+2, value=str("1"))   #симметрично заполняем таблицу относительно диагонали
				else:
					ws.cell(row=j + 2, column=k + 2, value=str("0"))
					ws.cell(row=k + 2, column=j + 2, value=str("0"))   #симметрично заполняем таблицу относительно диагонали
			elif tab == 3:
				var="round"
				ws.cell(row=j + 2, column=k + 2, value="{:.1f}".format(n))
				ws.cell(row=k + 2, column=j + 2, value="{:.1f}".format(n))
	name=str(mypath + "/Compare_{0}_{1}.xlsx".format(str(var), str(tol)))
	wb.save(name)
	#Начало сравнения осей
	wb = load_workbook(name)
	ws = wb.worksheets[0]
	lst = []	#список имен функционалов
	lst2=[]	# Создаем список сортировки
	for j in range(206):
		lst.append(ws.cell(row=1, column=j+2).value)
	print len(lst)
	for i in range(2, len(lst)+2): #+2 потому что так построен входной файл, цикл по строкам
		find=False
		if len(lst2)>0:
			for k in range(len(lst2)):
				if lst2[k].count(lst[i-2])!=0:
					find=True
		if find==True:
			continue
		lst2.append([])
		lst2[len(lst2)-1].append(lst[i-2])
		for j in range(2, len(lst)+2): #+2 потому что так построен входной файл, цикл по столбцам
			if i==j:
				continue
			if int(ws.cell(row=i, column=j).value)==1:
				#print "sovpalo"
				lst2[len(lst2)-1].append(lst[j-2]) #Добавляем элемент в список, если он отсутствует
				
			#else:
				#print  ws.cell(row=i, column=j).value
	pathCompareFile = str(mypath + "/Compare_{0}_{1}.txt".format(str(var), str(tol)))
	f = open(pathCompareFile, 'wt')
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Сравнение осей следа'
	for i in range(len(lst2)):
		print lst2[i]
		for j in range(len(lst2[i])):
			f.write(str(lst2[i][j]).replace('u', '').replace('\'','') + '\r\n')
			ws.cell(row=j+2, column=i+1, value="%s" % str(lst2[i][j]).replace('u', '').replace('\'',''))
	print len(lst2)
	name=str(mypath + "/Axis_{0}_{1}.xlsx".format(str(var), str(tol)))
	wb.save(name)
	f.close()
	
	
	return


if __name__ == "__main__":
	sys.exit(main())
