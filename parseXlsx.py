#!/usr/bin/python2.7
# -*- coding: utf-8 -*-
import sys
import os
import xml.etree.ElementTree as etree
from math import cos,sin,radians,sqrt,fabs,acos,pi,degrees,floor,log10
import xml.etree.ElementTree
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
import numpy as np
pathToAxis="/home/egor/work/CompareAll/"
pathToResult="/home/egor/work/results"

def find_element_in_list(element,list_element):
        try:
		index_element=list_element.index(element)
		return index_element
	except ValueError:
		return -1 

def find_substr_in_list(element,list_element):
	index_element = []
	for i, item in enumerate(list_element):
		if item.find(element) >= 0:
			index_element.append(i)
	return index_element

def readLineF(line,splitter = " "):
	lst = line.rstrip()
	vals = filter(None,lst.split(splitter))
	return [float(x) for x in vals]

def readLineI(line,splitter = " "):
	lst = line.rstrip()
	vals = filter(None,lst.split(splitter))
	return [int(x) for x in vals]

def cleanBlock(block):
	for i in range (len(block)-1, -1, -1):
		if(block[i].strip().strip("\t") == ""):
			del block[i]
	return block

def make_sure_path_exists(path):
	try: 
		os.makedirs(path)
	except OSError:
		if not os.path.isdir(path):
			raise

class GeoGrid:
	def __init__(self,gridFunction):
		self.lonmin = 0.0
		self.lonmax = 0.0
		self.latmin = 0.0
		self.latmax = 0.0
		self.countx = 0.0
		self.county = 0.0
		self.fmin   = 0.0
		self.fmax   = 0.0
		
		self.calcFunctionId = -1
		self.tipObl = -1
		self.isotop = ""
		self.phchForm = ""
		self.organ = ""
		self.age = ""
		self.timeEnd = 0.0
		
		self.data   = []

		if "calcFunctionId" not in gridFunction.attrib:
			self.calcFunctionId = -1
		else:
			self.calcFunctionId = int(gridFunction.get("calcFunctionId").strip())
		
		if "tipObl" not in gridFunction.attrib:
			self.tipObl = -1
		else:
			self.tipObl = int(gridFunction.get("tipObl").strip())
		
		if "isotop" not in gridFunction.attrib:
			self.isotop = ""
		else:
			self.isotop = gridFunction.get("isotop").strip()
			
		if "phchForm" not in gridFunction.attrib:
			self.phchForm = ""
		else:
			self.phchForm = gridFunction.get("phchForm").strip()
			
		if "organ" not in gridFunction.attrib:
			self.organ = ""
		else:
			self.organ = gridFunction.get("organ").strip()
			
		if "age" not in gridFunction.attrib:
			self.age = ""
		else:
			self.age = gridFunction.get("age").strip()
			
		if "timeEnd" not in gridFunction.attrib:
			self.timeEnd = 0.0
		else:
			self.timeEnd = float(gridFunction.get("timeEnd").strip())
				
		lst = gridFunction.text.split("\n")
		lst = cleanBlock(lst)

		self.countx = int(lst[0].split(" ")[1])
		self.county = int(lst[1].split(" ")[1])
		self.lonmin = float(lst[2].split(" ")[1])
		self.latmin = float(lst[3].split(" ")[1])
		self.dlon  =  float(lst[4].split(" ")[1])
		self.dlat  =  float(lst[5].split(" ")[1])
		self.lonmax = self.lonmin+self.dlon*(self.countx-1)
		self.latmax = self.latmin+self.dlat*(self.county-1)
		self.grap=np.array([[0.0]*self.county]*self.countx)
		lst = lst[7:]
		self.data   = [0.0]*self.countx*self.county
		for j in range(len(lst)): #количество строк
			lin = readLineF(lst[j])
			for i in range(len(lin)):
				self.data[i*self.county+(self.county-j-1)] = lin[i]
				self.grap[i][j]=lin[i]
		
		self.fmin   = self.findMin()
		self.fmax   = self.findMax()

		#self.printASCIIGRDFile("test.grd")
	
	def findMin(self):
		if(len(self.data) == 0):
			return -1
		ret = self.data[0]
		for v in self.data:
			if(ret > v):
				ret = v
		return ret
		
	def findMax(self):
		if(len(self.data) == 0):
			return -1
		ret = self.data[0]
		for v in self.data:
			if(ret < v):
				ret = v
		return ret			

def main():

	alltime = 366*24*3600/7200
	for ii in range(alltime+1):
		wb_new = openpyxl.Workbook()
		ws_new = wb_new.worksheets[0]
		ws_new.title = u'Сравнение осей следа'
		path = pathToAxis+"Axis_"+str(ii*7200)+".xlsx"
		print path
		wb = load_workbook(path)
		ws = wb.active
		lst=[]
		listOfFunc=[]
		et = etree.parse(str(pathToResult+"/"+str(ii*7200)+" s/out.xml"))
		root = et.getroot()
		grid = root.find('grid')
		
		for k, child_of_root in enumerate(root.iter()):
			if k>6:
				lst.append(child_of_root.attrib)
				fil = GeoGrid(child_of_root)
				print fil.calcFunctionId
		for i in range(1, int(ws.cell(row=2, column=2).value)+1):
			j=0
			#print "Axis   "+str(i)
			while True:
				try:
					str(ws.cell(row=3+j, column=i).value).split('f')[1]
				except:	
					break
				
				num = int(str(ws.cell(row=3+j, column=i).value).split('f')[1])
				ws_new.cell(row=3+j, column=i, value=str(lst[num]))
				
							#ws_new.cell(row=3+j, column=i, value=str('Tag: %s' % ( child_of_root.attrib)))
							#gridFunction.get("calcFunctionId").strip()
				j=j+1
		name=str(pathToAxis + "/Сomparison_Full/Compare_"+str(ii*7200)+"_Full.xlsx")
		wb_new.save(name)
		if ii >5:		
			break
		
		

	return 



if __name__ == "__main__":
	sys.exit(main())