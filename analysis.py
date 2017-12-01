#!/usr/bin/python2.7
# -*- coding: utf-8 -*-
#Модуль для анализа данных по различным авариям. Распределяет количество осей следа для каждой аварии 
#по массиву и выводит его в end.xlsx файл, в котором так же пишется процентное соотношение количества осей следа
#Запуск: 
#1. Запускается модуль createVariationalSeriesFI 
#2. Запускается getAxisForAll.py
#3. Запускается данный модуль
#Разработан: 27.11.2017 
#Автор: Ильичев Е.А.
#Последняя модификация: 27.11.2017 
import sys
import os
import xml.etree.ElementTree as etree
from math import cos,sin,radians,sqrt,fabs,acos,pi,degrees,floor,log10
import xml.etree.ElementTree
import openpyxl
from openpyxl import load_workbook
import numpy as np
path="/home/egor/work/CompareAll" # Папка где лежат файлы с именем Axis_*.xlsx, где *-время аварии



def main():
	lst=[]
	for i in range(207):
		lst.append([])	#создаем пустой список для имен начала аварии
		
	wb_new = openpyxl.Workbook()        #создаем новый xlsx файл
	ws_new = wb_new.worksheets[0]
	ws_new.title = u'Сравнение осей следа'
	
	numAxis=np.array([0.0]*207)
	alltime = 366*24*3600/7200
	for ii in range(alltime+1):
		name=str(path + "/Axis_"+str(ii*7200)+".xlsx")	#открываем файлы для считывания количества типов осей следа
		wb = load_workbook(name)
		ws = wb.active
		numAxis[int(ws.cell(row=2, column=2).value)+1]+=1	#заполняем массив, где индекс = количеству типов осей следа
		lst[int(ws.cell(row=2, column=2).value)+1].append(ii*7200)	#Заполняем список имен времени аварий
		#ws_new.cell(row=2, column=2, value=ii*7200)
		print ii*7200
		#print lst
		if ii*7200==24897600: # Последняя рассчитанная авария
			break
	summ=0.0
	ws_new.cell(row=1, column=1, value=u'Количество Осей Следа')
	ws_new.cell(row=2, column=1, value=u'Процент такого же количества от всех')
	for i in numAxis:
		summ+=i
	for i in range(len(lst)):
		ws_new.cell(row=1, column=2+i, value=i)
		if i>=1:
			ws_new.cell(row=2, column=i+1, value=numAxis[i]/summ*100.0)
		for j in range(len(lst[i])):
			
			ws_new.cell(row=3+j, column=1+i, value=lst[i][j])
	wb_new.save('end.xlsx')
	return 



if __name__ == "__main__":
	sys.exit(main())