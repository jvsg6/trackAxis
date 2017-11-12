#!/usr/bin/python2.7
# -*- coding: utf-8 -*-
#Модуль для сравнения осей следа.
#Реализует сравнение осей следа с помощью xlsx файла, созданного модулем createXlsxForCompareAxis.py 
#преобразование выходных данных ПС НОСТРА в более удобный для программной обработки формат. Строятся сетки, отражающие значения целевых функций,
#Запуск: 
#1. Запускается модуль createVariationalSeriesFI 
#2. Запускается createGRDFromOut.py 
#3. Запускается createTrackAxis.py
#4. Запускается createXlsxForCompareAxis.py
#5. Запускается данный модуль
#Разработан: 12.11.2017 
#Автор: Ильичев Е.А.
#Последняя модификация: 12.11.2017 
import sys
import os

from math import cos,sin,radians,sqrt,fabs,acos,pi,degrees,floor

import xml.etree.ElementTree
from geographiclib.geodesic import Geodesic
from geopy.distance import VincentyDistance
geod = Geodesic.WGS84
import numpy as np
import openpyxl
from openpyxl import load_workbook

pathToXlsxFile="/home/egor/work/Compare_01_10_95.xlsx"  # путь к xlsx файлу


def main():
	wb = load_workbook(pathToXlsxFile)
	ws = wb.worksheets[0]
	lst = []	#список имен функционалов
	lst2=[]	# Создаем список сортировки
	for j in range(206):
		lst.append(ws.cell(row=1, column=j+2).value)
	print len(lst)
	for i in range(2, len(lst)+2): #+2 потому что так построен входной файл, цикл по строкам
		if lst2.count(lst[i-2])!=0:    #проверяем, есть ли функционал в списке сортировки
			break
		lst2.append([])
		lst2[i-2].append(lst[i-2])
		for j in range(i, len(lst)+2): #+2 потому что так построен входной файл, цикл по столбцам
			if i==j:
				continue
			if int(ws.cell(row=i, column=j).value)==1:
				#print "sovpalo"
				lst2[i-2].append(lst[j-2]) #Добавляем элемент в список, если он отсутствует
				ws.cell(row=i, column=j).value=0
				
			#else:
				#print  ws.cell(row=i, column=j).value
		

	print lst2
	
	return


if __name__ == "__main__":
	sys.exit(main())
